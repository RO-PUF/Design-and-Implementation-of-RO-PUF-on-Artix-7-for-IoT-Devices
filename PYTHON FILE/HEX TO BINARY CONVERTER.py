from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from tqdm import tqdm  # Import tqdm for the progress bar

def hex_to_bin(hex_str: str) -> str:
    """
    Convert a hexadecimal string (no '0x' prefix) to a binary string,
    padded with leading zeros to a length of 4 * len(hex_str).
    """
    hex_str = hex_str.strip().upper()
    return bin(int(hex_str, 16))[2:].zfill(len(hex_str) * 4)

def convert_hex_range_to_binary(input_path: str, sheet_name: str = "Majority HEX"):
    # Load the workbook
    wb = load_workbook(input_path)

    # Check if the "Majority HEX" sheet exists
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
        return

    # Get the sheet by name
    ws = wb[sheet_name]

    # Create a new sheet called "Majority Binary"
    ws_binary = wb.create_sheet("Majority Binary")

    # Determine numeric index of column "T" (last column to process)
    max_col = column_index_from_string("T")

    # Wrap the row iteration in a tqdm progress bar to update periodically
    with tqdm(total=256 * max_col, desc="Processing Rows", unit="cell") as pbar:
        # Iterate rows 2 to 257, columns 2 (B) to max_col (T)
        for row in range(2, 258):  # Iterate over rows 2 to 257
            for col in range(1, max_col + 1):  # Iterate over columns 1 to max_col (T)
                cell = ws.cell(row=row, column=col)
                cell_binary = ws_binary.cell(row=row, column=col)  # Corresponding cell in the new sheet
                
                if cell.value is not None:
                    val = str(cell.value)
                    try:
                        # Convert hex to binary and store it in the new sheet
                        cell_binary.value = hex_to_bin(val)
                    except ValueError:
                        # If not a valid hex value, copy the value as-is
                        cell_binary.value = cell.value

                # Update the progress bar for each cell processed
                pbar.update(1)

    # Save the workbook with the new "Majority Binary" sheet
    wb.save(input_path)
    print(f"Conversion complete. Saved to:\n{input_path}")

if __name__ == "__main__":
    INPUT_FILE  = r"C:\Users\USER\Documents\vivado\XOR - INVERTER RO PUF\EVALUATION ANALYSIS\PLACEMENT CONFIGURATION DESIGN RESULTS\PLACEMENT CONFIGURATION DESIGN L.xlsx"
    convert_hex_range_to_binary(INPUT_FILE)
