import os
import itertools
from openpyxl import load_workbook, Workbook

# === CONFIGURATION ===
input_files = [
    rf"C:\ROPUF\FINAL RESULTS\RESULTS {chr(ord('A') + i)}.xlsx" for i in range(12)
]

output_dir = r"C:\ROPUF\FINAL RESULTS\COMBINED_RESULTS1"
FINAL_OUTPUT = r"C:\ROPUF\FINAL RESULTS\FINAL_RESULTS.xlsx"

os.makedirs(output_dir, exist_ok=True)

RESULT_FILES = [chr(ord('A') + i) for i in range(12)]  # A to L
all_combinations = list(itertools.combinations(range(12), 4))

# === HELPER FUNCTIONS ===
def copy_columns_limited_rows(source_ws, target_ws, start_col_idx, max_row):
    max_col = source_ws.max_column
    for src_col in range(1, max_col + 1):
        header_value = source_ws.cell(row=1, column=src_col).value
        label = header_value if header_value else f"Column {src_col}"
        target_ws.cell(row=1, column=start_col_idx).value = label
        for row_idx in range(2, max_row + 1):
            value = source_ws.cell(row=row_idx, column=src_col).value
            if value is not None and str(value).strip() != "":
                target_ws.cell(row=row_idx, column=start_col_idx).value = value
        start_col_idx += 1
    return start_col_idx

def hex_to_bin(hex_string):
    return bin(int(hex_string, 16))[2:].zfill(len(hex_string) * 4)

def hamming_distance(bin1, bin2):
    return sum(c1 != c2 for c1, c2 in zip(bin1, bin2))

def calculate_uniqueness(responses):
    bins = [hex_to_bin(r) for r in responses if r]
    k = len(bins)
    if k < 2:
        return 0.0
    n = len(bins[0])
    total_norm_hd = sum(hamming_distance(bins[i], bins[j]) / n for i in range(k) for j in range(i + 1, k))
    return (2 * total_norm_hd / (k * (k - 1))) * 100

def calculate_uniformity(responses):
    bins = [hex_to_bin(r) for r in responses if r]
    total_ones = sum(b.count('1') for b in bins)
    total_bits = len(bins) * len(bins[0])
    return (total_ones / total_bits) * 100 if total_bits else 0.0

def analyze_combination(filepath, start_row=2, end_row=257, show_bar=False):
    wb = load_workbook(filepath)
    if "Majority HEX" not in wb.sheetnames or "Reliability" not in wb.sheetnames:
        print(f"❌ {filepath} missing required sheets. Skipping...")
        return None
    sheet1 = wb["Majority HEX"]
    sheet2 = wb["Reliability"]

    groups = []
    max_col = sheet1.max_column
    total_rows = end_row - start_row + 1
    progress_counter = 0

    for row in sheet1.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col, values_only=True):
        vals = [cell for cell in row if cell not in (None, "")]
        if vals:
            groups.append(vals)
        progress_counter += 1
        if show_bar:
            show_progress(progress_counter, total_rows)

    uniquenesses = []
    uniformities = []
    counts = []

    for grp in groups:
        u = calculate_uniqueness(grp)
        v = calculate_uniformity(grp)
        uniquenesses.append(u)
        uniformities.append(v)
        counts.append(len(grp))

    total_samples = sum(counts)
    if total_samples:
        weighted_mean_uniqueness = sum(u * c for u, c in zip(uniquenesses, counts)) / total_samples
        weighted_mean_uniformity = sum(v * c for v, c in zip(uniformities, counts)) / total_samples
    else:
        weighted_mean_uniqueness = 0.0
        weighted_mean_uniformity = 0.0

    reliability_values = [cell[0] for cell in sheet2.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=1, values_only=True) if cell[0] is not None]
    weighted_mean_reliability = sum(reliability_values) / len(reliability_values) if reliability_values else 0.0

    return weighted_mean_uniqueness, weighted_mean_uniformity, weighted_mean_reliability

def show_progress(current, total, bar_length=30):
    fraction = current / total
    filled_length = int(bar_length * fraction)
    bar = '█' * filled_length + '▒' * (bar_length - filled_length)
    percent = fraction * 100
    print(f"\rProgress: [{bar}] {percent:.1f}%", end='')

# === STEP 1: Create 15 combinations ===
print("=== Combining files into 496 combinations ===\n")

for combo_indices in all_combinations:
    combo_letters = ''.join(RESULT_FILES[i] for i in combo_indices)
    output_filename = f"Combination {combo_letters}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    final_wb = Workbook()
    final_majority_ws = final_wb.active
    final_majority_ws.title = "Majority HEX"
    final_reliability_ws = final_wb.create_sheet("Reliability")

    majority_col_idx = 1
    reliability_col_idx = 1

    for i in combo_indices:
        file = input_files[i]
        if not os.path.exists(file):
            print(f"Warning: {file} not found. Skipping...")
            continue

        wb = load_workbook(file)
        if "Majority HEX" not in wb.sheetnames or "Reliability" not in wb.sheetnames:
            print(f"Warning: {file} missing required sheets. Skipping...")
            wb.close()
            continue

        source_majority = wb["Majority HEX"]
        source_reliability = wb["Reliability"]

        majority_col_idx = copy_columns_limited_rows(source_majority, final_majority_ws, majority_col_idx, max_row=257)
        reliability_col_idx = copy_columns_limited_rows(source_reliability, final_reliability_ws, reliability_col_idx, max_row=257)

        wb.close()

    final_wb.save(output_path)
    print(f"✅ Saved: {output_filename}")

print("\n🎯 All combinations successfully generated!")

# === STEP 2: Find best combination ===
print("\n=== Analyzing combinations for BEST uniqueness ===\n")

files = sorted([f for f in os.listdir(output_dir) if f.endswith('.xlsx')])

best_file = None
best_uniqueness = -1
best_combination_idx = None
best_result = None

for idx, file in enumerate(files, start=1):
    filepath = os.path.join(output_dir, file)
    result = analyze_combination(filepath, start_row=2, end_row=257)

    if result:
        wm_uniqueness, wm_uniformity, wm_reliability = result
        if wm_uniqueness > best_uniqueness:
            best_uniqueness = wm_uniqueness
            best_file = file
            best_combination_idx = idx - 1
            best_result = result

    show_progress(idx, len(files))

print()

# === STEP 3: Show the best combination and its parameters ===
if best_file is not None:
    selected_indices = all_combinations[best_combination_idx]
    print("\n🏆 Best Combination:")
    print(f"  Files Combined: {', '.join(RESULT_FILES[i] for i in selected_indices)}")
    print(f"  Best combination parameters:")
    print(f"    - Weighted mean uniqueness : {best_result[0]:.2f}%")
    print(f"    - Weighted mean uniformity : {best_result[1]:.2f}%")
    print(f"    - Weighted mean reliability: {best_result[2]:.2f}%")
else:
    print("\n❌ No valid combinations found.")
