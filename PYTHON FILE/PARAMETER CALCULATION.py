import os
from openpyxl import load_workbook, Workbook

# === CONFIGURATION ===
XLSX_PATH = r'C:\ROPUF\FINAL RESULTS\COMBINATION AIJK.xlsx'
START_ROW = 2
END_ROW = 257
SHEET1_NAME = "Majority HEX"
SHEET2_NAME = "Reliability"
OUTPUT_PATH = r'C:\ROPUF\FINAL RESULTS\FINAL RESULTS ANALYSIS (AIJK).xlsx'

# === HELPERS ===
def hex_to_bin(hex_string):
    return bin(int(hex_string, 16))[2:].zfill(len(hex_string) * 4)

def hamming_distance(a, b):
    return sum(x != y for x, y in zip(a, b))

def calculate_uniqueness(responses):
    bins = [hex_to_bin(r) for r in responses]
    k = len(bins)
    if k < 2:
        return 0.0
    n = len(bins[0])
    total = sum(
        hamming_distance(bins[i], bins[j]) / n
        for i in range(k) for j in range(i + 1, k)
    )
    return (2 * total / (k * (k - 1))) * 100

def calculate_uniformity(responses):
    bins = [hex_to_bin(r) for r in responses]
    ones = sum(b.count('1') for b in bins)
    bits = len(bins) * (len(bins[0]) if bins else 0)
    return (ones / bits) * 100 if bits else 0.0

def show_progress(current, total, length=30):
    frac = current / total
    filled = int(length * frac)
    bar = '█' * filled + '▒' * (length - filled)
    print(f'\r[{bar}] {frac*100:5.1f}% ({current}/{total})', end='', flush=True)

# === LOAD ===
if not os.path.exists(XLSX_PATH):
    raise FileNotFoundError(f"{XLSX_PATH} not found")

wb = load_workbook(XLSX_PATH, data_only=True)
maj = wb[SHEET1_NAME]
rel = wb[SHEET2_NAME]

# === GATHER ROWS WITH DATA ===
groups = []
rows = []
for ridx, row in enumerate(
    maj.iter_rows(min_row=START_ROW, max_row=END_ROW, values_only=True),
    start=START_ROW
):
    vals = [c for c in row if c not in (None, "")]
    if vals:
        groups.append(vals)
        rows.append(ridx)

# === COMPUTE METRICS ===
uniquenesses = []
uniformities = []
reliabilities = []
max_u, max_row = -1, None

print(f"\nComputing metrics for {len(groups)} challenges:")
for i, ridx in enumerate(rows, start=1):
    grp = groups[i-1]
    u = calculate_uniqueness(grp)
    v = calculate_uniformity(grp)
    # reliability for this same row
    r_val = rel.cell(row=ridx, column=1).value or 0.0

    uniquenesses.append(u)
    uniformities.append(v)
    reliabilities.append(r_val)

    if u > max_u:
        max_u, max_row = u, ridx

    show_progress(i, len(groups))
print()

# === SUMMARY STATS ===
counts = [len(g) for g in groups]
total_samples = sum(counts) or 1

wm_uni = sum(u*c for u, c in zip(uniquenesses, counts)) / total_samples
wm_unif = sum(v*c for v, c in zip(uniformities, counts)) / total_samples
wm_rel = sum(reliabilities) / len(reliabilities) if reliabilities else 0.0

# === WRITE TO EXCEL ===
out_wb = Workbook()

# Summary sheet
sum_ws = out_wb.active
sum_ws.title = "Summary"
sum_ws.append(["Metric", "Value"])
sum_ws.append(["Weighted mean uniqueness (%)", round(wm_uni, 2)])
sum_ws.append(["Weighted mean uniformity (%)", round(wm_unif, 2)])
sum_ws.append(["Average reliability (%)",         round(wm_rel, 2)])
sum_ws.append([
    "Challenge with highest uniqueness",
    f"{max_row-1} ({round(max_u,2)}%)"
])

# Details sheet
det_ws = out_wb.create_sheet("Details")
det_ws.append(["Challenge", "Uniqueness (%)", "Uniformity (%)", "Reliability (%)"])
for ridx, u, v, r in zip(rows, uniquenesses, uniformities, reliabilities):
    det_ws.append([ridx-1, round(u,2), round(v,2), round(r,2)])

out_wb.save(OUTPUT_PATH)
print(f"\n✅ All results saved to: {OUTPUT_PATH}")
