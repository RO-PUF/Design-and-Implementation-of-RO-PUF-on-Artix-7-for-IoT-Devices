import os
from openpyxl import load_workbook, Workbook

# === CONFIGURATION ===
input_files = [
    r"C:\ROPUF\FINAL RESULTS\RESULTS C.xlsx",
    r"C:\ROPUF\FINAL RESULTS\RESULTS F.xlsx",
    r"C:\ROPUF\FINAL RESULTS\RESULTS J.xlsx",
    r"C:\ROPUF\FINAL RESULTS\RESULTS K.xlsx"
]

output_file = r"C:\ROPUF\FINAL RESULTS\COMB1.xlsx"

# === CREATE FINAL WORKBOOK ===
final_wb = Workbook()
final_majority_ws = final_wb.active
final_majority_ws.title = "Majority HEX"
final_reliability_ws = final_wb.create_sheet("Reliability")

# === HELPER FUNCTION ===
def copy_all_columns(source_ws, target_ws, start_col_idx):
    """Copy all columns from source sheet to target sheet."""
    max_col = source_ws.max_column
    for src_col in range(1, max_col + 1):
        # Get header value
        header_value = source_ws.cell(row=1, column=src_col).value
        label = header_value if header_value else f"Column {src_col}"
        target_ws.cell(row=1, column=start_col_idx).value = label

        # Copy all rows (skip header row 1)
        for row_idx in range(2, source_ws.max_row + 1):
            value = source_ws.cell(row=row_idx, column=src_col).value
            if value is not None and str(value).strip() != "":
                target_ws.cell(row=row_idx, column=start_col_idx).value = value

        start_col_idx += 1

    return start_col_idx

# === COMBINE DATA ===
majority_col_idx = 1
reliability_col_idx = 1

for file in input_files:
    if not os.path.exists(file):
        print(f"Warning: {file} not found. Skipping...")
        continue

    wb = load_workbook(file)

    if "Majority HEX" not in wb.sheetnames or "Reliability" not in wb.sheetnames:
        print(f"Warning: {file} missing required sheets. Skipping...")
        continue

    source_majority = wb["Majority HEX"]
    source_reliability = wb["Reliability"]

    # Copy all columns from Majority HEX sheet
    majority_col_idx = copy_all_columns(source_majority, final_majority_ws, majority_col_idx)

    # Copy all columns from Reliability sheet
    reliability_col_idx = copy_all_columns(source_reliability, final_reliability_ws, reliability_col_idx)

    wb.close()

# Save final combined workbook
final_wb.save(output_file)
print("\n✅ Successfully combined ALL columns (without filenames) into:", output_file)
