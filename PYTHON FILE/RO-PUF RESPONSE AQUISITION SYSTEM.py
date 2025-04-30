import os
import time
import serial
import tkinter as tk
from tkinter import ttk, messagebox
import winsound
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from collections import Counter

# === CONFIGURATION ===
DEVICE_COM_PORT_MAP = {
    "F1": "COM18",
    "F2": "COM24",
    "F3": "COM8",
    "F4": "COM22",
    "F5": "COM20",
    "F6": "COM12", 
    "F7": "COM14",
    "F8": "COM4",
    "F9": "COM10",
    "F10": "COM16",
    "F11": "COM31",
    "F12": "COM19",
    "F13": "COM35",
    "F14": "COM37",
    "F15": "COM33",
    "F17": "COM41",
    "F16": "COM43",
    "F18": "COM29",    
    "F19": "COM30",
    "F20": "COM10"
}

RESULT_LETTERS = [chr(ord('A') + i) for i in range(12)]  # A to L

SAVE_DIR = r'C:\\ROPUF\\FINAL RESULTS'
os.makedirs(SAVE_DIR, exist_ok=True)

BAUD_RATE = 230400
TIMEOUT = 1
BATCH_SIZE = 128
MAX_ROWS = 256

LAST_CHOICE_FILE = os.path.join(SAVE_DIR, "last_choice.txt")

# === FUNCTIONS ===
def hex_to_bin(hex_str: str) -> str:
    return ''.join(f"{int(c, 16):04b}" for c in hex_str)

def bin_to_hex(bin_str: str) -> str:
    return ''.join(f"{int(bin_str[i:i+4], 2):X}" for i in range(0, len(bin_str), 4))

def compute_majority_bits(bin_list):
    transposed = zip(*bin_list)
    return ''.join(Counter(bits).most_common(1)[0][0] for bits in transposed)

def hamming_distance(bin1, bin2):
    return sum(b1 != b2 for b1, b2 in zip(bin1, bin2))

def next_empty_col(ws):
    col_idx = 1
    while ws.cell(row=2, column=col_idx).value not in (None, ""):
        col_idx += 1
    return col_idx

def save_last_choice(device, result_letter):
    with open(LAST_CHOICE_FILE, "w") as f:
        f.write(f"{device},{result_letter}")

def load_last_choice():
    if os.path.exists(LAST_CHOICE_FILE):
        with open(LAST_CHOICE_FILE, "r") as f:
            parts = f.read().strip().split(',')
            if len(parts) == 2:
                return parts[0], parts[1]
    return None, None

def play_alarm():
    duration = 10000  # milliseconds
    freq = 750  # Hz
    winsound.Beep(freq, duration)

# === MAIN LOGIC ===
def start_collection():
    device = device_var.get()
    result_letter = result_var.get()

    if device not in DEVICE_COM_PORT_MAP or result_letter not in RESULT_LETTERS:
        messagebox.showerror("Error", "Invalid device or result selection.")
        return

    save_last_choice(device, result_letter)

    HEADER_TEXT = f"{device} - {result_letter}"
    COM_PORT = DEVICE_COM_PORT_MAP[device]
    XLSX_FILE = os.path.join(SAVE_DIR, f"RESULTS {result_letter}.xlsx")

    # Open serial port
    try:
        ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=TIMEOUT)
    except Exception as e:
        messagebox.showerror("Serial Error", str(e))
        return

    # Setup workbook
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.create_sheet("Majority HEX")
        wb.create_sheet("Reliability")

    sheet1 = wb["Majority HEX"]
    sheet2 = wb["Reliability"]

    col1_idx = next_empty_col(sheet1)
    col2_idx = next_empty_col(sheet2)
    col1_letter = get_column_letter(col1_idx)
    col2_letter = get_column_letter(col2_idx)

    sheet1[f"{col1_letter}1"] = HEADER_TEXT
    sheet2[f"{col2_letter}1"] = HEADER_TEXT

    wb.save(XLSX_FILE)

    # Start main loop
    batch = []
    response_counter = 0
    start_time = time.time()

    progress_bar["maximum"] = MAX_ROWS
    progress_bar["value"] = 0

    try:
        while response_counter < MAX_ROWS:
            line = ser.readline().decode(errors='ignore').strip()

            if line and len(line) == 64:
                batch.append(line)
                if len(batch) == BATCH_SIZE:
                    bin_batch = [hex_to_bin(resp) for resp in batch]
                    majority_bin = compute_majority_bits(bin_batch)
                    majority_hex = bin_to_hex(majority_bin)

                    total_hd = sum(hamming_distance(majority_bin, b) for b in bin_batch)
                    avg_hd = total_hd / BATCH_SIZE
                    n_bits = len(majority_bin)
                    hd_intra_percent = (avg_hd / n_bits) * 100
                    reliability = 100 - hd_intra_percent

                    sheet1[f"{col1_letter}{response_counter + 2}"] = majority_hex
                    sheet2[f"{col2_letter}{response_counter + 2}"] = round(reliability, 2)

                    response_counter += 1
                    batch.clear()

                    progress_bar["value"] = response_counter
                    root.update_idletasks()

    except KeyboardInterrupt:
        print("\nInterrupted by user.")

    finally:
        wb.save(XLSX_FILE)
        ser.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        minutes = int(elapsed_time // 60)
        seconds = int(elapsed_time % 60)

        play_alarm()
        
        

# === GUI WINDOW ===
root = tk.Tk()
root.title("ROPUF Data Collection")

device_label = tk.Label(root, text="Select Device:")
device_label.pack()

device_var = tk.StringVar()
device_dropdown = ttk.Combobox(root, textvariable=device_var, values=list(DEVICE_COM_PORT_MAP.keys()), state="readonly")
device_dropdown.pack()

result_label = tk.Label(root, text="Select Result File:")
result_label.pack()

result_var = tk.StringVar()
result_dropdown = ttk.Combobox(root, textvariable=result_var, values=RESULT_LETTERS, state="readonly")
result_dropdown.pack()

start_button = tk.Button(root, text="Start Collection", command=start_collection)
start_button.pack(pady=10)

progress_bar = ttk.Progressbar(root, length=300)
progress_bar.pack(pady=10)

# Load last choice if available
last_device, last_result = load_last_choice()
if last_device and last_result:
    device_var.set(last_device)
    result_var.set(last_result)
else:
    device_var.set(list(DEVICE_COM_PORT_MAP.keys())[0])
    result_var.set(RESULT_LETTERS[0])

root.mainloop()